# PythonCastleHill

#Using Python to look for unexpected correlations in Closewool Ewe fertility Data

import os
import openpyxl
import csv
from datetime import date

#the following was found on web for getting icloud path. It didnt really work but was not needed so long as i put file in same folder as python script
#print(os.path.abspath('/Pythonista3/ClosewoolEwesAllBred.xlsx'))
file = open("ClosewoolEwesAllBred.csv",newline='')

reader = csv.reader(file)
header = next(reader)
#print(header)

wb = openpyxl.load_workbook("ClosewoolEwesAllBred.xlsx")
#print(wb.sheetnames)
#print(os.getcwd())
sheet=wb['Sheet1']
#print(sheet['a1'].value)

#create my output file in local folder
f = open('workfile.txt', 'w')

ugen=1 #not used
utag=0 #ewe's tag
ubday=0 #ewe's birthyear
ucountbyyear=[0,0,0,0,0]

f.write("Tag Number, birthyear, lambs in years 2016-2020\n")

#max row in spreadsheet is 1017
for i in range(2,1017,1):
	d = date.fromordinal(sheet.cell(row=i,column=5).value)
	uyear=d.year+1899 #a necessary tweak

	if (sheet.cell(row=i, column=1).value) == None:
		ugen=ugen+1
		
	else:
		print(utag, ubday, ucountbyyear)
		#next version will do this in single f.write
		f.write(str(utag)) 
		f.write(", ")
		f.write(str(ubday))
		f.write(", ")
		f.write(str(ucountbyyear))
		f.write("\n")
		for j in range(5): #there's probably a neater way to reset array
			ucountbyyear[j]=0
		utag = sheet.cell(row=i,column=1).value
		ubday = sheet.cell(row=i,column=2).value	
	
	for j in range(5):
		if uyear == 2016+j: #stores each year 2016-2020 in 5 entries
			ucountbyyear[j]=ucountbyyear[j]+1

f.close()

